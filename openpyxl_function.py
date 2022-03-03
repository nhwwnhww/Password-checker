import openpyxl

# create excel file
# book = openpyxl.Workbook()
# book.save(u'my_excel.xlsx')

# create sheet
# book = openpyxl.load_workbook(u'my_excel.xlsx')
# sheet = book.create_sheet('my_excel')
# book.save(u'my_excel.xlsx')
# book.close()

# load excel and write excel
# book = openpyxl.load_workbook(u'my_excel.xlsx')
# sheet = book['Sheet']
# sheet.cell(1,1).value = 'name'
# sheet.cell(2,1).value = 'book'
# sheet.cell(3,1).value = 'lol'
# book.save(u'my_excel.xlsx')
# book.close()

# get value
# book = openpyxl.load_workbook(u'my_excel.xlsx')
# sheet = book['Sheet']
# cell = sheet.cell(3,1)
# print('value:%s'%cell.value)